"""Временный скрипт: посмотреть содержимое Excel-шаблона с рабочего стола."""
import glob
import os

import openpyxl

desktop = os.path.join(os.path.expanduser("~"), "Desktop")
files = [f for f in glob.glob(os.path.join(desktop, "*.xlsx"))
         if not os.path.basename(f).startswith("~$")]
print("FILES FOUND:", files)

for path in files:
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        print("=== FILE:", path, "-> ERROR:", exc)
        continue
    print("=== FILE:", path)
    print("SHEETS:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print("--- sheet:", name, "| rows:", ws.max_row, "| cols:", ws.max_column)
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            print(i, row)
            if i >= 8:
                break

